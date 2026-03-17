import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os

os.makedirs("/home/ubuntu/rani_beauty", exist_ok=True)

wb = openpyxl.Workbook()

# ── Colour palette ──────────────────────────────────────────────────────────
ROSE_GOLD   = "C9956C"   # warm rose-gold
BLUSH       = "F5E6E0"   # light blush background
CREAM       = "FDF8F5"   # off-white row fill
DEEP_PLUM   = "4A2040"   # dark header text
SAGE        = "8FAF8A"   # accent green
SLATE       = "6B7280"   # secondary text
WHITE       = "FFFFFF"
LIGHT_GREY  = "F0EBEA"

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def row_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D1C4C0")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="2D2D2D"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

def apply_header_row(ws, row_num, values, fill_color=ROSE_GOLD, font_color=WHITE):
    for col_num, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=val)
        cell.fill = hdr_fill(fill_color)
        cell.font = hdr_font(color=font_color)
        cell.alignment = wrap_align("center", "center")
        cell.border = thin_border()

def apply_data_row(ws, row_num, values, fill_color=CREAM):
    for col_num, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=val)
        cell.fill = row_fill(fill_color)
        cell.font = body_font()
        cell.alignment = wrap_align()
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_and_filter(ws, cell="A3"):
    ws.freeze_panes = cell
    ws.auto_filter.ref = ws.dimensions

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER SHOT LIST
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Master Shot List"

# Title banner
ws1.merge_cells("A1:L1")
title_cell = ws1["A1"]
title_cell.value = "RANI BEAUTY CLINIC — MASTER PHOTOGRAPHY & VIDEO SHOT LIST"
title_cell.fill = hdr_fill(DEEP_PLUM)
title_cell.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32

# Column headers
headers = [
    "Shot #", "Category", "Shot Name / Description",
    "Shot Type", "Lens / Focal Length", "Lighting Setup",
    "Props & Styling", "Model / Talent", "Consent Required",
    "Platform Use", "Priority", "Notes / Director's Cue"
]
apply_header_row(ws1, 2, headers, fill_color=ROSE_GOLD)
ws1.row_dimensions[2].height = 28

# ── Shot data ───────────────────────────────────────────────────────────────
shots = [
    # ── CATEGORY 1: TREATMENT ROOM SETUP ──────────────────────────────────
    ("1.01", "Treatment Room Setup", "Wide establishing shot — full treatment room",
     "Wide / Establishing", "24–35 mm", "Soft box + natural window light",
     "Fresh white linens, rose-gold tray, branded candle, orchid", "None", "No",
     "Website Hero, Google Business", "High",
     "Shoot at golden hour for warm tones; ensure no clutter visible"),
    ("1.02", "Treatment Room Setup", "Treatment bed detail — pillow & draping",
     "Detail / Close-up", "85–100 mm macro", "Diffused natural light",
     "Monogrammed pillow, folded towel, small orchid sprig", "None", "No",
     "Instagram Feed, Website", "High",
     "Shoot from 45° angle; use shallow DOF to blur background"),
    ("1.03", "Treatment Room Setup", "Instrument tray flat-lay — sterile setup",
     "Overhead / Flat-lay", "50 mm", "Softbox overhead",
     "Branded tray liner, syringes (capped), vials, gauze, gloves", "None", "No",
     "Instagram, TikTok BTS", "High",
     "All items must be sealed/sterile; no blood or used items"),
    ("1.04", "Treatment Room Setup", "Consultation desk — provider side",
     "Medium / Environmental", "35 mm", "Practical desk lamp + fill light",
     "Laptop open to branded portal, notepad, pen, branded mug", "None", "No",
     "Website About Page, LinkedIn", "Medium",
     "Ensure laptop screen shows a tasteful, branded graphic"),
    ("1.05", "Treatment Room Setup", "Ambient mood shot — candle & botanicals",
     "Detail / Lifestyle", "85 mm", "Candlelight + subtle warm fill",
     "Branded candle, eucalyptus, rose petals, marble surface", "None", "No",
     "Instagram Stories, Pinterest", "Medium",
     "Shoot in portrait orientation for Stories; include brand color palette"),
    ("1.06", "Treatment Room Setup", "Skincare device on treatment cart",
     "Detail / Product", "100 mm macro", "Ring light at 45°",
     "Clean device, branded cloth underneath, small plant accent", "None", "No",
     "Website Services Page, Ads", "High",
     "Feature specific device per service (RF, laser, HydraFacial, etc.)"),
    ("1.07", "Treatment Room Setup", "Overhead bird's-eye of full room layout",
     "Overhead / Architectural", "16–24 mm wide", "All room lights on + fill",
     "Room fully dressed; remove personal items", "None", "No",
     "Website, Google Maps", "Medium",
     "Use a tripod arm or ladder; ensure symmetry"),

    # ── CATEGORY 2: PROVIDER ACTION SHOTS ────────────────────────────────
    ("2.01", "Provider Action", "Injector marking treatment zones on face",
     "Medium Close-up", "85 mm", "Softbox at 45° + reflector",
     "White coat, branded gloves, surgical marker", "Model (consented)", "Yes",
     "Website, Instagram, Ads", "High",
     "Capture focus on provider's eyes showing concentration"),
    ("2.02", "Provider Action", "Neuromodulator injection — side profile",
     "Close-up / Detail", "100 mm macro", "Diffused key light",
     "Syringe, gloved hands, clean skin prep", "Model (consented)", "Yes",
     "Instagram Reels, TikTok", "High",
     "Shoot multiple angles; ensure needle tip is in frame but not graphic"),
    ("2.03", "Provider Action", "Dermal filler — lip augmentation",
     "Close-up / Detail", "100 mm macro", "Ring light + fill",
     "Cannula or needle, lip balm nearby, gloved hands", "Model (consented)", "Yes",
     "Instagram, TikTok, Before/After", "High",
     "Capture before, during, and immediately after in sequence"),
    ("2.04", "Provider Action", "HydraFacial wand gliding over cheek",
     "Medium Close-up", "85 mm", "Natural window + reflector",
     "HydraFacial device, clean skin, branded headband on model", "Model (consented)", "Yes",
     "Instagram Reels, Website", "High",
     "Slow-motion capture at 120fps for Reels; show serum tip"),
    ("2.05", "Provider Action", "Microneedling pen on forehead",
     "Close-up / Detail", "100 mm", "Softbox overhead",
     "Microneedling device, numbing cream residue (if applicable)", "Model (consented)", "Yes",
     "TikTok, Instagram", "High",
     "Capture the motion stroke; avoid showing discomfort"),
    ("2.06", "Provider Action", "Chemical peel application with fan brush",
     "Medium / Action", "85 mm", "Softbox at 45°",
     "Fan brush, peel solution, branded bowl, gloves", "Model (consented)", "Yes",
     "Instagram, Website Services", "Medium",
     "Capture the brush stroke motion; use burst mode"),
    ("2.07", "Provider Action", "Laser device treatment — hand piece on skin",
     "Close-up / Detail", "100 mm macro", "Ambient + fill light",
     "Laser hand piece, protective eyewear on model", "Model (consented)", "Yes",
     "Instagram, TikTok, Website", "High",
     "Ensure protective eyewear is visible on both provider and model"),
    ("2.08", "Provider Action", "Provider reviewing skin analysis on screen",
     "Medium / Environmental", "35 mm", "Monitor glow + fill",
     "Skin analysis device, branded white coat, tablet", "Provider only", "No",
     "Website, LinkedIn, Email", "Medium",
     "Show provider pointing at screen with engaged expression"),
    ("2.09", "Provider Action", "Provider applying post-treatment serum",
     "Medium Close-up", "85 mm", "Soft natural light",
     "Branded serum, gloved hands, model with glowing skin", "Model (consented)", "Yes",
     "Instagram, Website Aftercare", "Medium",
     "Capture the gentle massage motion; golden-hour warmth preferred"),
    ("2.10", "Provider Action", "Provider and client laughing during consultation",
     "Medium / Candid", "50 mm", "Natural window light",
     "Notepad, tablet, branded coffee cups", "Model (consented)", "Yes",
     "Website About, Instagram Stories", "High",
     "Candid moment; do not over-direct — let it feel natural"),

    # ── CATEGORY 3: BEFORE / AFTER STAGING ───────────────────────────────
    ("3.01", "Before / After Staging", "Standardised face — front view B&A",
     "Portrait / Headshot", "85 mm", "Ring light or beauty dish — identical both shots",
     "Neutral background (grey or white), no makeup", "Model (consented)", "Yes",
     "Website, Instagram Carousel, Ads", "High",
     "CRITICAL: identical lighting, distance, expression both shots; use chin rest"),
    ("3.02", "Before / After Staging", "Standardised face — 45° angle B&A",
     "Portrait / 3/4", "85 mm", "Same as 3.01",
     "Same as 3.01", "Model (consented)", "Yes",
     "Website, Instagram Carousel", "High",
     "Mark floor position with tape for exact replication"),
    ("3.03", "Before / After Staging", "Standardised face — side profile B&A",
     "Portrait / Profile", "85 mm", "Same as 3.01",
     "Same as 3.01", "Model (consented)", "Yes",
     "Website, Instagram Carousel", "High",
     "Capture jawline and neck definition clearly"),
    ("3.04", "Before / After Staging", "Lip close-up B&A (filler)",
     "Macro / Detail", "100 mm macro", "Ring light — identical both shots",
     "Neutral lip, no gloss before; natural after", "Model (consented)", "Yes",
     "Instagram, TikTok, Website", "High",
     "Shoot at same distance; use lip liner as reference point"),
    ("3.05", "Before / After Staging", "Skin texture close-up B&A (HydraFacial / peel)",
     "Macro / Detail", "100 mm macro", "Raking side light to show texture",
     "No makeup both shots; same cheek area", "Model (consented)", "Yes",
     "Instagram, TikTok, Website", "High",
     "Side lighting reveals texture improvement dramatically"),
    ("3.06", "Before / After Staging", "Neck & décolletage B&A (laser / RF)",
     "Medium / Detail", "85 mm", "Diffused overhead",
     "Strapless top or drape; consistent framing", "Model (consented)", "Yes",
     "Website, Instagram", "Medium",
     "Include measurement reference if possible"),
    ("3.07", "Before / After Staging", "Hand rejuvenation B&A",
     "Detail / Overhead", "100 mm macro", "Flat overhead softbox",
     "Hands flat on white surface; no nail polish", "Model (consented)", "Yes",
     "Website, Instagram", "Low",
     "Ensure identical hand position and lighting angle"),

    # ── CATEGORY 4: PRODUCT FLAT-LAYS ────────────────────────────────────
    ("4.01", "Product Flat-lay", "Hero skincare product collection — brand palette",
     "Overhead / Flat-lay", "50 mm", "Overhead diffused softbox",
     "Marble surface, rose petals, eucalyptus, branded tissue", "None", "No",
     "Instagram Feed, Pinterest, Website Shop", "High",
     "Arrange in triangle composition; brand colors must be visible"),
    ("4.02", "Product Flat-lay", "Single hero product — serum bottle close-up",
     "Detail / Product", "100 mm macro", "Side window light + reflector",
     "Dewy water droplets on bottle, minimal props", "None", "No",
     "Instagram, Website Product Page", "High",
     "Use a spray bottle for water droplets; shoot multiple angles"),
    ("4.03", "Product Flat-lay", "Monthly treatment kit — curated collection",
     "Overhead / Flat-lay", "50 mm", "Overhead softbox",
     "Branded box, tissue paper, products arranged neatly", "None", "No",
     "Instagram, Email Newsletter", "Medium",
     "Capture both open box and closed box versions"),
    ("4.04", "Product Flat-lay", "Pre-treatment prep kit flat-lay",
     "Overhead / Flat-lay", "50 mm", "Overhead softbox",
     "Cleanser, toner, SPF, headband, branded bag", "None", "No",
     "Instagram Stories, Website", "Medium",
     "Label each item with text overlay in post-production"),
    ("4.05", "Product Flat-lay", "Post-treatment recovery kit flat-lay",
     "Overhead / Flat-lay", "50 mm", "Overhead softbox",
     "Healing balm, SPF, cooling mask, branded instructions card", "None", "No",
     "Instagram, Website Aftercare", "Medium",
     "Include a handwritten-style note card for warmth"),
    ("4.06", "Product Flat-lay", "Injectables product lineup (branded vials)",
     "Overhead / Detail", "100 mm macro", "Side softbox",
     "Branded vials, syringes (capped), medical tray liner", "None", "No",
     "Website, Instagram (educational)", "High",
     "Ensure all vials are sealed; no patient-identifiable information"),
    ("4.07", "Product Flat-lay", "Gift set / retail display styling",
     "Medium / 3/4 angle", "50 mm", "Natural window + reflector",
     "Branded gift box, ribbon, tissue, seasonal accent (flowers)", "None", "No",
     "Instagram, Email, Holiday Campaigns", "Medium",
     "Shoot seasonal variants: spring florals, winter pine, etc."),

    # ── CATEGORY 5: STAFF HEADSHOTS ───────────────────────────────────────
    ("5.01", "Staff Headshots", "Lead provider — formal headshot (white coat)",
     "Portrait / Headshot", "85 mm", "Beauty dish + fill reflector",
     "White coat, stethoscope optional, branded badge", "Provider", "No",
     "Website Team Page, LinkedIn, Press", "High",
     "Shoot against neutral grey or blush backdrop; 3 expressions"),
    ("5.02", "Staff Headshots", "Lead provider — approachable lifestyle headshot",
     "Portrait / Environmental", "50 mm", "Natural window light",
     "Smart casual outfit, treatment room background (blurred)", "Provider", "No",
     "Instagram Bio, Google Business, Email", "High",
     "Warm, genuine smile; capture mid-laugh variant"),
    ("5.03", "Staff Headshots", "Full team — formal group portrait",
     "Group / Wide", "35 mm", "Multiple softboxes for even coverage",
     "Matching white coats or coordinated outfits", "All Staff", "No",
     "Website About Page, Press Kit", "High",
     "Arrange by height; ensure all faces are clearly lit"),
    ("5.04", "Staff Headshots", "Full team — candid group lifestyle",
     "Group / Candid", "35 mm", "Natural light + reflector",
     "Casual branded attire, clinic interior background", "All Staff", "No",
     "Instagram, Website About", "High",
     "Capture genuine interaction; avoid stiff poses"),
    ("5.05", "Staff Headshots", "Individual staff — formal headshots (all team members)",
     "Portrait / Headshot", "85 mm", "Beauty dish + fill",
     "White coat, branded badge", "Each Staff Member", "No",
     "Website Team Page, LinkedIn", "High",
     "Shoot 3 expressions per person: serious, smiling, candid"),
    ("5.06", "Staff Headshots", "Provider hands detail — artisan shot",
     "Detail / Close-up", "100 mm macro", "Side window light",
     "Clean gloved hands holding syringe or device", "Provider", "No",
     "Website, Instagram (provider spotlight)", "Medium",
     "Focus on the craft and precision of the hands"),
    ("5.07", "Staff Headshots", "Provider in consultation — over-shoulder shot",
     "Medium / Environmental", "50 mm", "Natural window light",
     "Tablet, notepad, branded pen", "Provider + Model", "Yes",
     "Website, Instagram Stories", "Medium",
     "Shoot from behind provider to show client perspective"),

    # ── CATEGORY 6: CLINIC EXTERIOR & INTERIOR ────────────────────────────
    ("6.01", "Clinic Exterior & Interior", "Exterior facade — daytime",
     "Wide / Architectural", "24 mm", "Natural daylight",
     "Clean signage, potted plants at entrance, branded door mat", "None", "No",
     "Google Business, Website, Ads", "High",
     "Shoot on overcast day for even light; avoid harsh shadows"),
    ("6.02", "Clinic Exterior & Interior", "Exterior facade — golden hour / evening",
     "Wide / Architectural", "24 mm", "Ambient golden light + interior glow",
     "Interior lights on, welcoming warm glow through windows", "None", "No",
     "Instagram, Google Business", "Medium",
     "Shoot 30 min before sunset; capture the inviting glow"),
    ("6.03", "Clinic Exterior & Interior", "Reception / front desk — wide shot",
     "Wide / Interior", "24–35 mm", "Overhead ambient + fill flash",
     "Fresh flowers, branded signage, clean desk, iPad check-in", "Receptionist", "No",
     "Website, Google Business, Instagram", "High",
     "Ensure branded elements are prominent; no personal items visible"),
    ("6.04", "Clinic Exterior & Interior", "Waiting area — lifestyle ambiance",
     "Medium / Interior", "35 mm", "Natural window + ambient",
     "Magazine, branded water bottle, fresh flowers, artwork", "None", "No",
     "Website, Instagram, Google", "High",
     "Capture the luxury and calm; shoot from client's seated perspective"),
    ("6.05", "Clinic Exterior & Interior", "Hallway / corridor leading to treatment rooms",
     "Wide / Architectural", "24 mm", "Overhead ambient",
     "Artwork on walls, branded signage, clean floors", "None", "No",
     "Website, Instagram Stories", "Medium",
     "Use leading lines of corridor for visual depth"),
    ("6.06", "Clinic Exterior & Interior", "Branded signage / logo wall close-up",
     "Detail / Close-up", "85 mm", "Side light to show dimension",
     "Logo wall, branded elements, accent lighting", "None", "No",
     "Instagram, Website, Press Kit", "High",
     "Capture both straight-on and angled perspectives"),
    ("6.07", "Clinic Exterior & Interior", "Retail display / product shelf",
     "Medium / Interior", "50 mm", "Natural + fill",
     "Neatly arranged products, branded shelf labels, accent lighting", "None", "No",
     "Website Shop, Instagram, Google", "Medium",
     "Shoot multiple angles; include a close-up of hero products"),
    ("6.08", "Clinic Exterior & Interior", "Bathroom / amenity area detail",
     "Detail / Interior", "35 mm", "Ambient + fill",
     "Branded hand soap, towels, fresh flowers, candle", "None", "No",
     "Instagram Stories, Website", "Low",
     "Highlight luxury touches that differentiate the experience"),

    # ── CATEGORY 7: CLIENT EXPERIENCE JOURNEY ────────────────────────────
    ("7.01", "Client Experience Journey", "Client arriving at entrance — door pull",
     "Medium / Action", "50 mm", "Natural daylight",
     "Client in stylish outfit, branded door, clean exterior", "Model (consented)", "Yes",
     "Instagram Reels, Website", "High",
     "Shoot from inside looking out; capture anticipation on face"),
    ("7.02", "Client Experience Journey", "Client checking in at reception",
     "Medium / Action", "50 mm", "Ambient + fill",
     "iPad check-in, receptionist smiling, branded desk", "Model + Receptionist", "Yes",
     "Instagram Reels, Website Journey", "High",
     "Capture warm welcome moment; both faces should be visible"),
    ("7.03", "Client Experience Journey", "Client in waiting area — relaxed",
     "Medium / Lifestyle", "50 mm", "Natural window light",
     "Branded water, magazine, client in comfortable pose", "Model (consented)", "Yes",
     "Instagram, Website", "Medium",
     "Capture calm, anticipatory mood; no phone in hand"),
    ("7.04", "Client Experience Journey", "Consultation — provider and client face-to-face",
     "Medium / Two-shot", "50 mm", "Natural window + fill",
     "Tablet, notepad, branded coffee, warm lighting", "Model + Provider", "Yes",
     "Website, Instagram, Ads", "High",
     "Show engaged listening; provider leaning slightly forward"),
    ("7.05", "Client Experience Journey", "Client reviewing treatment plan on tablet",
     "Medium Close-up", "85 mm", "Natural light",
     "Branded tablet cover, client nodding/smiling", "Model (consented)", "Yes",
     "Website, Instagram", "Medium",
     "Capture the informed-consent moment; show confidence"),
    ("7.06", "Client Experience Journey", "Client reclining in treatment chair — pre-treatment",
     "Medium / Lifestyle", "50 mm", "Soft ambient",
     "Branded headband, robe, relaxed expression", "Model (consented)", "Yes",
     "Instagram, Website", "High",
     "Capture the luxury recline; show comfort and trust"),
    ("7.07", "Client Experience Journey", "Treatment in progress — provider and client",
     "Medium / Action", "50 mm", "Softbox + natural",
     "Treatment device, provider in white coat, client relaxed", "Model + Provider", "Yes",
     "Instagram Reels, TikTok, Website", "High",
     "Capture the care and precision; client should appear comfortable"),
    ("7.08", "Client Experience Journey", "Post-treatment glow — client looking in mirror",
     "Medium / Lifestyle", "50 mm", "Natural window light",
     "Branded mirror, client touching face gently, glowing skin", "Model (consented)", "Yes",
     "Instagram, TikTok, Website", "High",
     "Capture the 'wow' reaction; shoot from behind and front"),
    ("7.09", "Client Experience Journey", "Aftercare instruction handoff",
     "Medium / Action", "50 mm", "Natural + fill",
     "Branded aftercare card, recovery kit bag, provider handing to client", "Model + Provider", "Yes",
     "Website, Instagram", "Medium",
     "Show the care that extends beyond the treatment room"),
    ("7.10", "Client Experience Journey", "Client departure — happy exit",
     "Medium / Action", "50 mm", "Natural daylight",
     "Client with branded bag, smiling at door, waving goodbye", "Model (consented)", "Yes",
     "Instagram Reels, Website, Ads", "High",
     "Capture the satisfaction and confidence of the transformed client"),
]

# Write data rows
for i, shot in enumerate(shots):
    fill = CREAM if i % 2 == 0 else BLUSH
    apply_data_row(ws1, i + 3, shot, fill_color=fill)

# Set column widths
set_col_widths(ws1, [7, 22, 38, 20, 20, 28, 38, 20, 14, 28, 10, 42])
ws1.row_dimensions[1].height = 32
for r in range(3, len(shots) + 3):
    ws1.row_dimensions[r].height = 52

freeze_and_filter(ws1, "A3")

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — SHOOT DAY SCHEDULE
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Shoot Day Schedule")

ws2.merge_cells("A1:H1")
t2 = ws2["A1"]
t2.value = "RANI BEAUTY CLINIC — SHOOT DAY SCHEDULE TEMPLATE"
t2.fill = hdr_fill(DEEP_PLUM)
t2.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

sched_headers = ["Time Slot", "Duration", "Location / Set", "Shot Numbers", "Talent Required", "Props Needed", "Photographer Notes", "Status"]
apply_header_row(ws2, 2, sched_headers, fill_color=ROSE_GOLD)

schedule = [
    ("7:00 – 7:30 AM", "30 min", "Clinic Exterior", "6.01", "None", "Clean signage, plants", "Shoot before foot traffic; golden morning light", "☐ To Do"),
    ("7:30 – 8:00 AM", "30 min", "Reception / Waiting Area", "6.03, 6.04, 6.05, 6.06", "Receptionist", "Fresh flowers, branded items", "Style reception before staff arrive", "☐ To Do"),
    ("8:00 – 9:00 AM", "60 min", "Treatment Room 1", "1.01–1.07", "None", "Full room styling kit", "Complete all room setup shots before models arrive", "☐ To Do"),
    ("9:00 – 9:30 AM", "30 min", "Product / Flat-lay Station", "4.01–4.07", "None", "All product inventory + props", "Use portable table with marble contact paper", "☐ To Do"),
    ("9:30 – 10:30 AM", "60 min", "Treatment Room 1 — Headshots", "5.01–5.07", "All Staff", "White coats, badges", "Individual then group shots; 3 expressions each", "☐ To Do"),
    ("10:30 – 11:00 AM", "30 min", "Break + Model Arrival", "—", "Models (consented)", "Consent forms, touch-up kit", "Review consent forms; brief models on poses", "☐ To Do"),
    ("11:00 AM – 12:30 PM", "90 min", "Treatment Room 1 — Action Shots", "2.01–2.10", "Lead Provider + Model 1", "Treatment props, devices", "Capture all provider action shots; rotate devices", "☐ To Do"),
    ("12:30 – 1:00 PM", "30 min", "Lunch Break", "—", "All", "—", "Review morning shots; adjust afternoon plan if needed", "☐ To Do"),
    ("1:00 – 2:00 PM", "60 min", "Before / After Station", "3.01–3.07", "Models 1 & 2", "Grey/white backdrop, chin rest, ring light", "Shoot all B&A angles; mark floor positions with tape", "☐ To Do"),
    ("2:00 – 3:30 PM", "90 min", "Client Journey Sequence", "7.01–7.10", "Model 2 + Receptionist + Provider", "Full journey props", "Shoot in chronological order; capture candid moments", "☐ To Do"),
    ("3:30 – 4:00 PM", "30 min", "Clinic Exterior — Golden Hour Setup", "6.02", "None", "Interior lights on", "Prepare for golden hour; set interior lighting", "☐ To Do"),
    ("4:00 – 4:30 PM", "30 min", "Clinic Exterior — Golden Hour Shoot", "6.02", "Optional: Provider at door", "Branded door mat, plants", "Shoot exterior with warm evening glow", "☐ To Do"),
    ("4:30 – 5:00 PM", "30 min", "Overflow / Reshoot Buffer", "TBD", "TBD", "TBD", "Use this time for any missed or retake shots", "☐ To Do"),
    ("5:00 PM", "—", "Wrap + Pack Down", "—", "All", "—", "Back up all footage; complete shot list checklist", "☐ To Do"),
]

for i, row in enumerate(schedule):
    fill = CREAM if i % 2 == 0 else BLUSH
    apply_data_row(ws2, i + 3, row, fill_color=fill)
    ws2.row_dimensions[i + 3].height = 40

set_col_widths(ws2, [18, 12, 28, 20, 28, 32, 42, 12])
ws2.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — CONSENT & MODEL TRACKER
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Consent & Model Tracker")

ws3.merge_cells("A1:J1")
t3 = ws3["A1"]
t3.value = "RANI BEAUTY CLINIC — MODEL CONSENT & RELEASE TRACKER"
t3.fill = hdr_fill(DEEP_PLUM)
t3.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

consent_headers = [
    "Model ID", "Full Name", "Role", "Shots Assigned",
    "Written Consent", "Photo Release", "Video Release",
    "Social Media Use", "Paid / Gifted", "Notes"
]
apply_header_row(ws3, 2, consent_headers, fill_color=ROSE_GOLD)

consent_rows = [
    ("M-001", "[Name]", "Treatment Model", "2.01–2.10, 3.01–3.07, 7.04–7.10",
     "☐ Signed", "☐ Signed", "☐ Signed", "☐ Approved", "Gifted treatment", "Before/after model — full face consent"),
    ("M-002", "[Name]", "Journey Model", "7.01–7.10",
     "☐ Signed", "☐ Signed", "☐ Signed", "☐ Approved", "Gifted treatment", "Client journey sequence; full body consent"),
    ("M-003", "[Name]", "Lip Filler Model", "2.03, 3.04",
     "☐ Signed", "☐ Signed", "☐ Signed", "☐ Approved", "Gifted filler", "Lip close-up only; no full face required"),
    ("S-001", "[Provider Name]", "Lead Provider", "2.01–2.10, 5.01–5.02, 7.04–7.09",
     "☐ Signed", "☐ Signed", "☐ Signed", "☐ Approved", "Staff", "Provider action shots and headshots"),
    ("S-002", "[Staff Name]", "Receptionist", "6.03, 7.02",
     "☐ Signed", "☐ Signed", "☐ Signed", "☐ Approved", "Staff", "Reception and check-in scenes"),
    ("S-003", "[Staff Name]", "Support Staff", "5.03–5.05",
     "☐ Signed", "☐ Signed", "☐ Signed", "☐ Approved", "Staff", "Group and individual headshots only"),
]

for i, row in enumerate(consent_rows):
    fill = CREAM if i % 2 == 0 else BLUSH
    apply_data_row(ws3, i + 3, row, fill_color=fill)
    ws3.row_dimensions[i + 3].height = 44

set_col_widths(ws3, [10, 20, 20, 32, 14, 14, 14, 14, 16, 36])
ws3.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — PROPS & STYLING CHECKLIST
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Props & Styling Checklist")

ws4.merge_cells("A1:F1")
t4 = ws4["A1"]
t4.value = "RANI BEAUTY CLINIC — PROPS & STYLING MASTER CHECKLIST"
t4.fill = hdr_fill(DEEP_PLUM)
t4.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 32

props_headers = ["Category", "Item", "Quantity", "Source / Supplier", "Prepared By", "Status"]
apply_header_row(ws4, 2, props_headers, fill_color=ROSE_GOLD)

props_data = [
    # Branded Items
    ("Branded Items", "White lab coats (branded logo)", "3", "Custom embroidery supplier", "Clinic Manager", "☐ Ready"),
    ("Branded Items", "Branded tray liners / paper", "10", "Print supplier", "Clinic Manager", "☐ Ready"),
    ("Branded Items", "Branded gift bags", "5", "Packaging supplier", "Clinic Manager", "☐ Ready"),
    ("Branded Items", "Branded candles", "3", "Candle supplier", "Clinic Manager", "☐ Ready"),
    ("Branded Items", "Branded water bottles / cups", "6", "Branded merchandise", "Clinic Manager", "☐ Ready"),
    ("Branded Items", "Branded aftercare instruction cards", "10", "Print supplier", "Clinic Manager", "☐ Ready"),
    ("Branded Items", "Branded notepad and pen set", "3", "Stationery supplier", "Clinic Manager", "☐ Ready"),
    ("Branded Items", "Branded headbands for models", "4", "Branded merchandise", "Clinic Manager", "☐ Ready"),
    # Medical Props
    ("Medical Props", "Capped syringes (unused, sealed)", "10", "Medical supplier", "Lead Provider", "☐ Ready"),
    ("Medical Props", "Sealed vials (branded or generic)", "6", "Medical supplier", "Lead Provider", "☐ Ready"),
    ("Medical Props", "Sterile gloves (various sizes)", "2 boxes", "Medical supplier", "Lead Provider", "☐ Ready"),
    ("Medical Props", "Surgical marker / skin pen", "3", "Medical supplier", "Lead Provider", "☐ Ready"),
    ("Medical Props", "Gauze pads (sealed)", "1 pack", "Medical supplier", "Lead Provider", "☐ Ready"),
    ("Medical Props", "Medical tray (stainless steel)", "2", "Medical supplier", "Lead Provider", "☐ Ready"),
    # Décor & Styling
    ("Décor & Styling", "Fresh white orchids (potted)", "4", "Florist", "Stylist", "☐ Ready"),
    ("Décor & Styling", "Eucalyptus bundles", "3", "Florist / market", "Stylist", "☐ Ready"),
    ("Décor & Styling", "Rose petals (loose)", "2 bags", "Florist", "Stylist", "☐ Ready"),
    ("Décor & Styling", "Marble contact paper (for flat-lay surface)", "1 roll", "Craft store", "Stylist", "☐ Ready"),
    ("Décor & Styling", "White linen set (treatment bed)", "3 sets", "Linen supplier", "Stylist", "☐ Ready"),
    ("Décor & Styling", "Neutral grey / white backdrop (portable)", "1", "Photography supplier", "Photographer", "☐ Ready"),
    ("Décor & Styling", "Chin rest / posing guide", "1", "Photography supplier", "Photographer", "☐ Ready"),
    ("Décor & Styling", "Floor position tape (coloured)", "1 roll", "Hardware store", "Photographer", "☐ Ready"),
    # Skincare Products
    ("Skincare Products", "Hero serum (for flat-lay and action shots)", "3 units", "Clinic inventory", "Clinic Manager", "☐ Ready"),
    ("Skincare Products", "SPF product (for flat-lay)", "2 units", "Clinic inventory", "Clinic Manager", "☐ Ready"),
    ("Skincare Products", "Cleanser (for flat-lay)", "2 units", "Clinic inventory", "Clinic Manager", "☐ Ready"),
    ("Skincare Products", "Post-treatment healing balm", "2 units", "Clinic inventory", "Clinic Manager", "☐ Ready"),
    ("Skincare Products", "Cooling gel mask", "2 units", "Clinic inventory", "Clinic Manager", "☐ Ready"),
    # Equipment
    ("Equipment", "Ring light (18-inch minimum)", "1", "Photography supplier", "Photographer", "☐ Ready"),
    ("Equipment", "Portable softbox set (2-light)", "1 set", "Photography supplier", "Photographer", "☐ Ready"),
    ("Equipment", "Reflector (5-in-1, 32-inch)", "1", "Photography supplier", "Photographer", "☐ Ready"),
    ("Equipment", "Tripod (heavy duty)", "1", "Photography supplier", "Photographer", "☐ Ready"),
    ("Equipment", "Memory cards (64GB+, x4)", "4", "Electronics store", "Photographer", "☐ Ready"),
    ("Equipment", "Laptop for tethered shooting", "1", "Photographer's own", "Photographer", "☐ Ready"),
]

for i, row in enumerate(props_data):
    fill = CREAM if i % 2 == 0 else BLUSH
    apply_data_row(ws4, i + 3, row, fill_color=fill)
    ws4.row_dimensions[i + 3].height = 36

set_col_widths(ws4, [20, 40, 12, 28, 20, 12])
ws4.freeze_panes = "A3"

# Save
output_path = "/home/ubuntu/rani_beauty/Rani_Beauty_Clinic_Shot_List.xlsx"
wb.save(output_path)
print(f"Shot list saved to: {output_path}")
