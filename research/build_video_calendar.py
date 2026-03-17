import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
DEEP_PLUM   = "4A2040"
ROSE_GOLD   = "C9956C"
BLUSH       = "F5E6E0"
CREAM       = "FDF8F5"
SAGE        = "8FAF8A"
SAGE_DARK   = "5A8A55"
LAVENDER    = "E8E0F0"
PEACH       = "FAE5D3"
MINT        = "D5EDD5"
LIGHT_BLUE  = "D5E8F5"
YELLOW      = "FFF8D5"
WHITE       = "FFFFFF"

SERVICE_COLORS = {
    "Neuromodulators (Botox/Dysport)": "F5D5E8",
    "Dermal Fillers": "F5E5D5",
    "HydraFacial": "D5EEF5",
    "Chemical Peels": "EEF5D5",
    "Laser Treatments": "F5F0D5",
    "Microneedling": "E8D5F5",
    "RF Skin Tightening": "D5F5EE",
    "IV Therapy / Wellness": "F5D5D5",
    "Lash & Brow Services": "D5D5F5",
    "Body Contouring": "F5EED5",
}

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="D1C4C0")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="2D2D2D"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def wrap_align(h="left", v="top"):
    return Alignment(wrap_text=True, horizontal=h, vertical=v)

def apply_header_row(ws, row_num, values, fill_color=ROSE_GOLD, font_color="FFFFFF"):
    for col_num, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=val)
        cell.fill = hdr_fill(fill_color)
        cell.font = hdr_font(color=font_color)
        cell.alignment = wrap_align("center", "center")
        cell.border = thin_border()

def apply_data_row(ws, row_num, values, fill_color=CREAM):
    for col_num, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_num, value=val)
        cell.fill = hdr_fill(fill_color)
        cell.font = body_font()
        cell.alignment = wrap_align()
        cell.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — 12-WEEK VIDEO CONTENT CALENDAR
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "12-Week Content Calendar"

# Title
ws1.merge_cells("A1:N1")
t1 = ws1["A1"]
t1.value = "RANI BEAUTY CLINIC — 12-WEEK REELS & TIKTOK VIDEO CONTENT CALENDAR"
t1.fill = hdr_fill(DEEP_PLUM)
t1.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
t1.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 34

# Sub-header
ws1.merge_cells("A2:N2")
s2 = ws1["A2"]
s2.value = "Posting cadence: 3–4 videos per week | Platforms: Instagram Reels + TikTok | Repurpose top performers to YouTube Shorts"
s2.fill = hdr_fill(BLUSH)
s2.font = Font(name="Calibri", size=10, italic=True, color=DEEP_PLUM)
s2.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 22

# Column headers
cal_headers = [
    "Week", "Post Date\n(Suggested)", "Service / Topic",
    "Concept #", "Concept Name", "Video Title / Hook",
    "Key Visual", "Audio / Sound Direction",
    "Caption Direction", "CTA", "Platform",
    "Consent Needed", "Shot List Refs", "Status"
]
apply_header_row(ws1, 3, cal_headers, fill_color=ROSE_GOLD)
ws1.row_dimensions[3].height = 40

# ── Calendar data ────────────────────────────────────────────────────────────
# 12 weeks × 3–4 posts per week = 40 entries
calendar_data = [
    # WEEK 1 — Neuromodulators
    (1, "Mon, Apr 7", "Neuromodulators", "#3 — The 'Why'",
     "Provider Education", '"3 reasons Botox is still the #1 treatment we do"',
     "Provider talking to camera in white coat, treatment room background",
     "Direct audio (lavalier mic) + subtle lo-fi background",
     "Educational tone. Explain the 3 benefits: prevention, correction, confidence. End with: 'Book a free consultation.'",
     "Link in bio to book", "Instagram Reels + TikTok", "No", "2.08, 5.01", "☐ To Film"),
    (1, "Wed, Apr 9", "Neuromodulators", "#9 — Myth vs. Fact",
     "Myth Busting", '"Botox makes you look frozen? Let\'s debunk that."',
     "Provider on camera with text overlays popping up",
     "Trending audio with text pop-up sound effects",
     "Use 3 myths. Keep it punchy and confident. End with: 'We specialize in natural results.'",
     "Save to 'Botox Facts' highlight", "Instagram Reels + TikTok", "No", "5.01, 5.02", "☐ To Film"),
    (1, "Fri, Apr 11", "Neuromodulators", "#4 — Treatment in Action",
     "Cinematic BTS", '"Watch the precision that goes into every treatment"',
     "Slow-motion close-up of injection, provider's focused expression",
     "Trending cinematic instrumental",
     "Short caption: 'Precision. Care. Results. Every time.' + treatment details.",
     "Book via link in bio", "Instagram Reels + TikTok", "Yes", "2.01, 2.02", "☐ To Film"),

    # WEEK 2 — Dermal Fillers
    (2, "Mon, Apr 14", "Dermal Fillers", "#5 — What to Expect",
     "Step-by-Step Guide", '"Your lip filler appointment, from start to finish"',
     "Text overlays on B-roll of each step: consult → numb → treat → reveal",
     "Upbeat trending sound",
     "Walk through the 5 steps. Reassure first-timers. End: 'Questions? DM us!'",
     "DM for questions", "Instagram Reels + TikTok", "Yes", "7.04–7.08", "☐ To Film"),
    (2, "Wed, Apr 16", "Dermal Fillers", "#6 — Before & After",
     "Transformation Reveal", '"The subtle lip enhancement that changed everything"',
     "Split-screen or smooth transition from before to after",
     "Trending transformation audio",
     "Let the results speak. Caption: 'Natural, beautiful, you — just enhanced.' Tag model (with consent).",
     "Book a consultation", "Instagram Reels + TikTok", "Yes", "3.01–3.04", "☐ To Film"),
    (2, "Fri, Apr 18", "Dermal Fillers", "#2 — BTS Prep",
     "ASMR / Detail", '"The satisfying setup before your filler appointment"',
     "Close-up ASMR: gloves snapping, vial opening, tray setup",
     "Natural ASMR sounds only — no music",
     "No text needed. Let the visuals and sounds do the work. Caption: 'The art of preparation.'",
     "Save to 'BTS' highlight", "TikTok + Instagram Reels", "No", "1.03, 2.01", "☐ To Film"),

    # WEEK 3 — HydraFacial
    (3, "Mon, Apr 21", "HydraFacial", "#1 — The Walk-Through",
     "POV Experience", '"Come with me to my HydraFacial appointment at Rani"',
     "POV: arriving at clinic, checking in, settling into treatment chair",
     "Calming luxury lo-fi beat",
     "First-person narration style. 'This is what a HydraFacial appointment actually looks like.' End: 'Book yours today.'",
     "Book via link in bio", "Instagram Reels + TikTok", "Yes", "7.01–7.06", "☐ To Film"),
    (3, "Wed, Apr 23", "HydraFacial", "#4 — Treatment in Action",
     "Cinematic Slow-Mo", '"Watching the HydraFacial wand work is oddly satisfying"',
     "Slow-motion of wand gliding over cheek, serum being extracted",
     "Trending satisfying/ASMR audio",
     "Caption: 'Deep cleanse. Hydration. Glow. All in one treatment.' + service details.",
     "Book a HydraFacial", "Instagram Reels + TikTok", "Yes", "2.04", "☐ To Film"),
    (3, "Fri, Apr 25", "HydraFacial", "#7 — Post-Treatment Glow",
     "Reaction / Reveal", '"Her reaction after her first HydraFacial said it all"',
     "Client looking in mirror, touching face, glowing skin close-up",
     "Upbeat, feel-good trending audio",
     "Let the client's reaction be the star. Caption: 'This is why we do what we do.' Tag with permission.",
     "Save to 'Results' highlight", "Instagram Reels + TikTok", "Yes", "7.08", "☐ To Film"),

    # WEEK 4 — Chemical Peels
    (4, "Mon, Apr 28", "Chemical Peels", "#3 — The 'Why'",
     "Provider Education", '"Is a chemical peel right for you? Here\'s what to know"',
     "Provider on camera, split with B-roll of peel application",
     "Direct audio + subtle background music",
     "Explain 3 skin concerns peels address. Mention downtime honestly. End: 'Book a skin consultation.'",
     "Book a skin consult", "Instagram Reels + TikTok", "No", "2.06, 5.01", "☐ To Film"),
    (4, "Wed, Apr 30", "Chemical Peels", "#8 — Aftercare Essentials",
     "Flat-lay / Stop Motion", '"Everything you need for your post-peel recovery"',
     "Stop-motion of products being placed into branded recovery bag",
     "Crisp, rhythmic beat",
     "Name each product and why it matters. Caption: 'We take care of you, even after you leave.'",
     "Shop aftercare products", "Instagram Reels + TikTok", "No", "4.04, 4.05", "☐ To Film"),
    (4, "Fri, May 2", "Chemical Peels", "#6 — Before & After",
     "Skin Texture Reveal", '"The skin texture transformation after 3 chemical peels"',
     "Macro close-up transition: rough texture before → smooth skin after",
     "Trending transformation audio",
     "Caption: 'Consistency is key. 3 peels, 3 months, incredible results.' Include treatment details.",
     "Book a peel series", "Instagram Reels + TikTok", "Yes", "3.05", "☐ To Film"),

    # WEEK 5 — Laser Treatments
    (5, "Mon, May 5", "Laser Treatments", "#5 — What to Expect",
     "Step-by-Step", '"Your first laser treatment: what actually happens"',
     "Text overlays on B-roll of each step",
     "Upbeat, clear audio",
     "Demystify laser treatments. Address the 'does it hurt?' question. End: 'It's more comfortable than you think.'",
     "Book a laser consult", "Instagram Reels + TikTok", "Yes", "2.07, 7.04–7.08", "☐ To Film"),
    (5, "Wed, May 7", "Laser Treatments", "#4 — Treatment in Action",
     "Cinematic BTS", '"The technology behind flawless skin"',
     "Close-up of laser hand piece on skin, protective eyewear visible",
     "Futuristic/tech-inspired trending audio",
     "Caption: 'Advanced technology. Proven results. Zero compromise.' + device name.",
     "Learn more on website", "Instagram Reels + TikTok", "Yes", "2.07", "☐ To Film"),
    (5, "Fri, May 9", "Laser Treatments", "#9 — Myth vs. Fact",
     "Myth Busting", '"Laser treatments are only for one skin type? Myth."',
     "Provider on camera, confident delivery",
     "Trending audio with text pop-ups",
     "Address the top 3 laser myths. Emphasize safety and inclusivity of skin tones.",
     "DM for skin type questions", "Instagram Reels + TikTok", "No", "5.01, 5.02", "☐ To Film"),

    # WEEK 6 — Microneedling
    (6, "Mon, May 12", "Microneedling", "#2 — BTS Prep",
     "ASMR / Detail", '"The satisfying setup before your microneedling session"',
     "ASMR: device unboxing, numbing cream application, tray setup",
     "Natural ASMR sounds only",
     "No narration needed. Caption: 'Precision. Preparation. Perfection.'",
     "Save to 'BTS' highlight", "TikTok + Instagram Reels", "No", "1.03, 2.05", "☐ To Film"),
    (6, "Wed, May 14", "Microneedling", "#3 — The 'Why'",
     "Provider Education", '"Why microneedling is our most requested skin treatment"',
     "Provider talking head + B-roll of treatment",
     "Direct audio + subtle background",
     "Explain collagen induction, who it's for, and realistic results timeline.",
     "Book a consultation", "Instagram Reels + TikTok", "No", "5.01", "☐ To Film"),
    (6, "Fri, May 16", "Microneedling", "#6 — Before & After",
     "Skin Transformation", '"6 weeks post-microneedling. The results speak for themselves."',
     "Macro skin close-up transition: before → after",
     "Trending transformation audio",
     "Caption: 'Collagen. Clarity. Confidence. This is what 6 weeks of healing looks like.'",
     "Book a microneedling series", "Instagram Reels + TikTok", "Yes", "3.05", "☐ To Film"),

    # WEEK 7 — RF Skin Tightening
    (7, "Mon, May 19", "RF Skin Tightening", "#3 — The 'Why'",
     "Provider Education", '"The non-surgical lift that actually works"',
     "Provider on camera, before/after images in background",
     "Direct audio + soft background music",
     "Explain how RF works, ideal candidates, and realistic expectations. No hype.",
     "Book a skin consult", "Instagram Reels + TikTok", "No", "5.01", "☐ To Film"),
    (7, "Wed, May 21", "RF Skin Tightening", "#4 — Treatment in Action",
     "Cinematic BTS", '"Watching RF tightening work in real time"',
     "Close-up of RF device on jawline/neck, client relaxed",
     "Calming, sophisticated trending audio",
     "Caption: 'Lift. Tighten. Contour. No surgery required.' + treatment details.",
     "Book RF treatment", "Instagram Reels + TikTok", "Yes", "2.07", "☐ To Film"),
    (7, "Fri, May 23", "RF Skin Tightening", "#6 — Before & After",
     "Jawline Reveal", '"The jawline definition she always wanted — no surgery"',
     "Side profile B&A transition",
     "Trending reveal audio",
     "Caption: 'Natural lift, natural results. This is what RF skin tightening can do for you.'",
     "Book a consultation", "Instagram Reels + TikTok", "Yes", "3.02, 3.03", "☐ To Film"),

    # WEEK 8 — IV Therapy / Wellness
    (8, "Mon, May 26", "IV Therapy / Wellness", "#1 — The Walk-Through",
     "POV Experience", '"What an IV therapy session at Rani actually looks like"',
     "POV: arriving, getting set up, relaxing during treatment",
     "Calming, spa-like lo-fi audio",
     "Normalize IV therapy. Show the comfort and luxury of the experience.",
     "Book IV therapy", "Instagram Reels + TikTok", "Yes", "7.01–7.07", "☐ To Film"),
    (8, "Wed, May 28", "IV Therapy / Wellness", "#3 — The 'Why'",
     "Provider Education", '"Why high-performers are adding IV therapy to their routine"',
     "Provider on camera, lifestyle B-roll of active, glowing clients",
     "Energetic, aspirational audio",
     "Target the wellness-focused client. Discuss energy, immunity, skin benefits.",
     "Book a wellness consult", "Instagram Reels + TikTok", "No", "5.01", "☐ To Film"),
    (8, "Fri, May 30", "IV Therapy / Wellness", "#10 — Day in the Life",
     "Vlog Style", '"A day at Rani: behind the scenes of our IV lounge"',
     "Fast-paced montage of multiple clients, setup, and provider interactions",
     "Energetic, fast-paced trending audio",
     "Show the community and luxury of the IV lounge. Caption: 'Your wellness, elevated.'",
     "Book a session", "Instagram Reels + TikTok", "Yes", "7.01–7.10", "☐ To Film"),

    # WEEK 9 — Lash & Brow Services
    (9, "Mon, Jun 2", "Lash & Brow Services", "#4 — Treatment in Action",
     "Cinematic Detail", '"The art of the perfect brow — watch this"',
     "Extreme close-up of brow mapping, tinting, or lamination process",
     "Satisfying, rhythmic trending audio",
     "Caption: 'Brows that frame your face. Precision is everything.'",
     "Book brow appointment", "Instagram Reels + TikTok", "Yes", "2.06", "☐ To Film"),
    (9, "Wed, Jun 4", "Lash & Brow Services", "#7 — Post-Treatment Glow",
     "Reaction / Reveal", '"Her reaction to her new brows made our day"',
     "Client looking in mirror, touching brows, smiling",
     "Feel-good trending audio",
     "Caption: 'The right brows change everything.' Tag client with permission.",
     "Save to 'Results' highlight", "Instagram Reels + TikTok", "Yes", "7.08", "☐ To Film"),
    (9, "Fri, Jun 6", "Lash & Brow Services", "#9 — Myth vs. Fact",
     "Myth Busting", '"Brow lamination ruins your brows? Let\'s talk about that."',
     "Provider on camera, confident and informative",
     "Trending audio with text pop-ups",
     "Address top 3 brow myths. Emphasize the safety and expertise at Rani.",
     "DM for brow questions", "Instagram Reels + TikTok", "No", "5.01", "☐ To Film"),

    # WEEK 10 — Body Contouring
    (10, "Mon, Jun 9", "Body Contouring", "#3 — The 'Why'",
     "Provider Education", '"Is body contouring right for you? The honest answer"',
     "Provider on camera, direct and trustworthy",
     "Direct audio + subtle background",
     "Be honest about expectations. Discuss ideal candidates and realistic results.",
     "Book a body consult", "Instagram Reels + TikTok", "No", "5.01", "☐ To Film"),
    (10, "Wed, Jun 11", "Body Contouring", "#5 — What to Expect",
     "Step-by-Step", '"Your body contouring appointment, step by step"',
     "Text overlays on B-roll of each step",
     "Upbeat, clear audio",
     "Walk through the process. Address common concerns about comfort and downtime.",
     "Book a consultation", "Instagram Reels + TikTok", "Yes", "7.04–7.08", "☐ To Film"),
    (10, "Fri, Jun 13", "Body Contouring", "#6 — Before & After",
     "Body Transformation", '"The results after a full body contouring series"',
     "B&A of targeted area (with full written consent)",
     "Trending transformation audio",
     "Caption: 'Consistency + the right treatment = results that last.' Include series details.",
     "Book a series", "Instagram Reels + TikTok", "Yes", "3.06", "☐ To Film"),

    # WEEK 11 — Clinic Culture / Brand
    (11, "Mon, Jun 16", "Clinic Culture", "#10 — Day in the Life",
     "Vlog Style", '"A day in the life at Rani Beauty Clinic"',
     "Fast-paced montage of a full clinic day: team, treatments, clients",
     "Energetic, upbeat trending audio",
     "Show the team's passion and the clinic's energy. Caption: 'This is why we love what we do.'",
     "Follow for more BTS", "Instagram Reels + TikTok", "Yes", "5.03, 5.04, 7.01–7.10", "☐ To Film"),
    (11, "Wed, Jun 18", "Clinic Culture", "#3 — The 'Why' (Brand Story)",
     "Brand Storytelling", '"Why we started Rani Beauty Clinic"',
     "Provider/founder on camera, personal and authentic",
     "Soft, emotional background music",
     "Share the 'why' behind the clinic. Connect emotionally with the audience.",
     "Save to 'Our Story' highlight", "Instagram Reels + TikTok", "No", "5.01, 5.02", "☐ To Film"),
    (11, "Fri, Jun 20", "Clinic Culture", "#1 — The Walk-Through",
     "Clinic Tour", '"Tour our clinic — your new favourite place"',
     "Smooth walk-through of entire clinic: exterior → reception → treatment rooms",
     "Luxury, aspirational lo-fi audio",
     "Invite viewers in. Caption: 'This is where your transformation begins.'",
     "Book a visit", "Instagram Reels + TikTok", "No", "6.01–6.08", "☐ To Film"),

    # WEEK 12 — Seasonal Campaign / Recap
    (12, "Mon, Jun 23", "Seasonal Campaign", "#6 — Before & After Compilation",
     "Results Compilation", '"3 months of transformations at Rani Beauty Clinic"',
     "Quick-cut compilation of best B&A results from the quarter",
     "Upbeat, celebratory trending audio",
     "Caption: '3 months. Countless transformations. This is Rani.' Tag all consenting clients.",
     "Book before summer", "Instagram Reels + TikTok", "Yes", "3.01–3.07", "☐ To Film"),
    (12, "Wed, Jun 25", "Seasonal Campaign", "#7 — Client Testimonials",
     "Social Proof / UGC", '"What our clients are saying about Rani"',
     "Compilation of UGC clips + client testimonials (with consent)",
     "Warm, feel-good audio",
     "Let clients speak. Minimal text. Caption: 'The best reviews come from our community.'",
     "Read more reviews", "Instagram Reels + TikTok", "Yes", "7.08, 7.10", "☐ To Film"),
    (12, "Fri, Jun 27", "Seasonal Campaign", "#3 — Summer Skin Prep",
     "Seasonal Education", '"The 3 treatments to book before summer"',
     "Provider on camera + B-roll of each treatment",
     "Upbeat, summer-inspired audio",
     "Timely, relevant content. Drive urgency for summer bookings.",
     "Book summer treatments", "Instagram Reels + TikTok", "No", "5.01, 2.04, 2.06, 2.07", "☐ To Film"),
]

for i, row in enumerate(calendar_data):
    fill = CREAM if i % 2 == 0 else BLUSH
    apply_data_row(ws1, i + 4, row, fill_color=fill)
    ws1.row_dimensions[i + 4].height = 60

set_col_widths(ws1, [7, 16, 22, 18, 22, 38, 38, 30, 44, 22, 22, 14, 14, 12])
ws1.freeze_panes = "A4"
ws1.auto_filter.ref = ws1.dimensions

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — VIDEO CONCEPT LIBRARY (10 Concepts × 10 Services = 100 entries)
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Video Concept Library")

ws2.merge_cells("A1:K1")
t2 = ws2["A1"]
t2.value = "RANI BEAUTY CLINIC — VIDEO CONCEPT LIBRARY (10 Concepts × 10 Services)"
t2.fill = hdr_fill(DEEP_PLUM)
t2.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 34

lib_headers = [
    "Service", "Concept #", "Concept Name",
    "Suggested Hook / Title", "Key Visual Description",
    "Duration", "Consent", "Platform Priority",
    "Content Pillar", "Difficulty", "Notes"
]
apply_header_row(ws2, 2, lib_headers, fill_color=ROSE_GOLD)
ws2.row_dimensions[2].height = 36

services = list(SERVICE_COLORS.keys())

concepts = [
    ("#1", "The Walk-Through", "Come with me to my [service] at Rani", "POV arrival → check-in → treatment chair", "30–45s", "Yes", "Instagram Reels", "Experience", "Medium", "Film in one continuous take if possible"),
    ("#2", "BTS Prep (ASMR)", "The satisfying setup before your [service]", "Close-up ASMR: gloves, tray, products", "15–30s", "No", "TikTok", "Behind-the-Scenes", "Easy", "No music — natural sounds only"),
    ("#3", "The 'Why' (Education)", "Why [service] is our most requested treatment", "Provider talking head + B-roll", "30–60s", "No", "Instagram Reels + TikTok", "Education", "Easy", "Lavalier mic essential"),
    ("#4", "Treatment in Action", "Watch the precision of [service] in slow-mo", "Cinematic slow-motion of treatment", "15–30s", "Yes", "Instagram Reels + TikTok", "Showcase", "Hard", "120fps slow-mo required"),
    ("#5", "What to Expect", "Your [service] appointment, step by step", "Text overlays on B-roll of each step", "30–45s", "Yes", "Instagram Reels", "Education", "Medium", "Pre-plan text overlays before filming"),
    ("#6", "Before & After", "The [service] transformation that changed everything", "Split-screen or smooth transition", "15–30s", "Yes", "Instagram Reels + TikTok", "Social Proof", "Medium", "Standardised lighting CRITICAL"),
    ("#7", "Post-Treatment Glow", "Her reaction after [service] made our day", "Client mirror reaction + skin close-up", "15–30s", "Yes", "Instagram Reels + TikTok", "Social Proof", "Easy", "Capture in the moment — don't over-direct"),
    ("#8", "Aftercare Essentials", "Everything you need after your [service]", "Stop-motion flat-lay of recovery products", "15–30s", "No", "Instagram Reels", "Education", "Medium", "Use branded products and packaging"),
    ("#9", "Myth vs. Fact", "The biggest [service] myths, debunked", "Provider on camera + text pop-ups", "30–45s", "No", "TikTok + Instagram Reels", "Education", "Easy", "Research top 3 myths for each service"),
    ("#10", "Day in the Life", "A day performing [service] at Rani", "Fast-paced vlog montage", "30–60s", "Yes", "Instagram Reels + TikTok", "Brand Culture", "Hard", "Film across multiple appointments for variety"),
]

row_num = 3
for service in services:
    svc_color = SERVICE_COLORS[service]
    for concept in concepts:
        row_data = [service] + list(concept)
        for col_num, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_num, column=col_num, value=val)
            cell.fill = hdr_fill(svc_color)
            cell.font = body_font()
            cell.alignment = wrap_align()
            cell.border = thin_border()
        ws2.row_dimensions[row_num].height = 44
        row_num += 1

set_col_widths(ws2, [28, 10, 22, 40, 40, 10, 10, 24, 18, 10, 36])
ws2.freeze_panes = "A3"
ws2.auto_filter.ref = ws2.dimensions

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — UGC STRATEGY TRACKER
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("UGC Strategy Tracker")

ws3.merge_cells("A1:I1")
t3 = ws3["A1"]
t3.value = "RANI BEAUTY CLINIC — UGC COLLECTION & MANAGEMENT TRACKER"
t3.fill = hdr_fill(DEEP_PLUM)
t3.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 34

ugc_headers = [
    "Client Name / ID", "Date of Visit", "Service Received",
    "UGC Type", "Platform Posted", "Tagged Clinic?",
    "Consent to Repost", "Reposted On", "Notes"
]
apply_header_row(ws3, 2, ugc_headers, fill_color=ROSE_GOLD)

ugc_template = [
    ("[Client Name]", "[Date]", "[Service]", "Photo / Video / Story / Reel",
     "Instagram / TikTok / Google", "☐ Yes  ☐ No",
     "☐ Verbal  ☐ Written  ☐ DM Reply", "[Date reposted / platform]", "[Notes]"),
] * 20

for i, row in enumerate(ugc_template):
    fill = CREAM if i % 2 == 0 else BLUSH
    apply_data_row(ws3, i + 3, row, fill_color=fill)
    ws3.row_dimensions[i + 3].height = 36

set_col_widths(ws3, [22, 14, 22, 28, 24, 16, 28, 24, 28])
ws3.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — EQUIPMENT CHECKLIST
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Equipment Checklist")

ws4.merge_cells("A1:G1")
t4 = ws4["A1"]
t4.value = "RANI BEAUTY CLINIC — SELF-PRODUCED CONTENT EQUIPMENT CHECKLIST"
t4.fill = hdr_fill(DEEP_PLUM)
t4.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 34

eq_headers = ["Category", "Item", "Recommended Model", "Approx. Price (USD)", "Priority", "Use Case", "Status"]
apply_header_row(ws4, 2, eq_headers, fill_color=ROSE_GOLD)

equipment = [
    # Camera
    ("Camera", "Primary Smartphone", "Apple iPhone 15 Pro / Samsung Galaxy S24 Ultra", "$999–$1,199", "Essential", "Daily Reels, TikToks, Stories, B&A photos", "☐ To Purchase"),
    ("Camera", "Dedicated Content Camera", "Sony ZV-E10 II / Panasonic Lumix G100", "$749–$899", "Recommended", "Higher quality video for hero content and website", "☐ To Purchase"),
    ("Camera", "Standard Lens", "Sony 16–50mm kit lens (or equivalent)", "Included", "Essential", "Wide shots, room tours, group shots", "☐ To Purchase"),
    ("Camera", "Portrait Lens", "Sony 35mm f/1.8 or 50mm f/1.8", "$249–$399", "Recommended", "Headshots, detail shots, shallow DOF", "☐ To Purchase"),
    # Lighting
    ("Lighting", "Large Ring Light (18-inch)", "Neewer 18-inch Bi-Color Ring Light Kit", "$89–$129", "Essential", "Talking heads, B&A photos, flat-lays", "☐ To Purchase"),
    ("Lighting", "LED Panel (Key Light)", "Elgato Key Light Air / Lume Cube Panel", "$99–$149", "Essential", "Treatment room fill light, product shots", "☐ To Purchase"),
    ("Lighting", "LED Panel (Fill Light)", "Neewer 2-Pack LED Panel Kit", "$79–$99", "Recommended", "Even lighting for multi-angle shoots", "☐ To Purchase"),
    ("Lighting", "5-in-1 Reflector (32-inch)", "Neewer 32-inch 5-in-1 Collapsible Reflector", "$19–$29", "Essential", "Bouncing natural light, reducing shadows", "☐ To Purchase"),
    # Audio
    ("Audio", "Wireless Lavalier Mic System", "DJI Mic 2 / Rode Wireless GO II", "$249–$299", "Essential", "All talking-head videos, provider education content", "☐ To Purchase"),
    ("Audio", "Shotgun Microphone", "Rode VideoMicro II", "$79–$99", "Recommended", "On-camera audio for B-roll and action shots", "☐ To Purchase"),
    # Stabilization
    ("Stabilization", "Heavy-Duty Tripod", "Joby GorillaPod 5K / Manfrotto Compact", "$79–$129", "Essential", "All stationary shots, B&A photos, flat-lays", "☐ To Purchase"),
    ("Stabilization", "Smartphone Gimbal", "DJI Osmo Mobile 6 / Zhiyun Smooth 5S", "$119–$149", "Recommended", "Walk-through videos, smooth tracking shots", "☐ To Purchase"),
    ("Stabilization", "Overhead Arm / Articulating Arm", "JOBY GorillaPod Arm or similar", "$39–$59", "Recommended", "Overhead flat-lay shots, product videos", "☐ To Purchase"),
    # Accessories
    ("Accessories", "Smartphone Lens Clip Set", "Moment Lens / Xenvo Pro Lens Kit", "$49–$99", "Optional", "Macro shots, wider angles on smartphone", "☐ To Purchase"),
    ("Accessories", "Memory Cards (V30, 64GB+)", "SanDisk Extreme Pro 64GB (×4)", "$19–$29 each", "Essential", "All camera and video storage", "☐ To Purchase"),
    ("Accessories", "Portable Power Bank", "Anker 20,000mAh PowerCore", "$39–$49", "Essential", "Keeping devices charged during long shoot days", "☐ To Purchase"),
    ("Accessories", "Flat-lay Surface (Marble Contact Paper)", "Con-Tact Brand Marble Adhesive Paper", "$12–$18", "Essential", "Product flat-lays and detail shots", "☐ To Purchase"),
    ("Accessories", "Portable Backdrop (White/Grey)", "Neewer 5×7ft Muslin Backdrop + Stand", "$49–$79", "Recommended", "Standardised B&A photos, headshots", "☐ To Purchase"),
    # Software
    ("Software / Apps", "Video Editing App (Mobile)", "CapCut (Free) / Adobe Premiere Rush ($9.99/mo)", "$0–$9.99/mo", "Essential", "Editing Reels and TikToks on-device", "☐ To Set Up"),
    ("Software / Apps", "Photo Editing App", "Lightroom Mobile (Free) / VSCO ($19.99/yr)", "$0–$19.99/yr", "Essential", "Consistent photo editing and brand presets", "☐ To Set Up"),
    ("Software / Apps", "Content Scheduling Tool", "Later / Planoly / Buffer ($15–$18/mo)", "$15–$18/mo", "Recommended", "Scheduling posts, planning the content calendar", "☐ To Set Up"),
    ("Software / Apps", "AI Image Generation", "Midjourney ($10/mo) / Adobe Firefly (included in CC)", "$0–$10/mo", "Optional", "Generating brand-aligned stock photo alternatives", "☐ To Set Up"),
]

for i, row in enumerate(equipment):
    fill = CREAM if i % 2 == 0 else BLUSH
    apply_data_row(ws4, i + 3, row, fill_color=fill)
    ws4.row_dimensions[i + 3].height = 40

set_col_widths(ws4, [18, 36, 40, 20, 14, 42, 14])
ws4.freeze_panes = "A3"

# Save
output_path = "/home/ubuntu/rani_beauty/Rani_Beauty_Clinic_Video_Content_Calendar.xlsx"
wb.save(output_path)
print(f"Video content calendar saved to: {output_path}")
