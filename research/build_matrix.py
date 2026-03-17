import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
import json

# ── colour palette ──────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"
SUBHDR_BG   = "2E75B6"   # mid-blue
SUBHDR_FG   = "FFFFFF"
ROW_ALT1    = "EBF3FB"   # light blue
ROW_ALT2    = "FFFFFF"
ACCENT_GRN  = "E2EFDA"   # light green  (positive)
ACCENT_RED  = "FFE7E7"   # light red    (weakness)
ACCENT_YLW  = "FFF2CC"   # yellow       (neutral / note)
TITLE_BG    = "C9DAF8"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(bold=True, size=11, color=HEADER_FG):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def cell_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

# ── data ─────────────────────────────────────────────────────────────────────
competitors = [
    {
        "id": 1,
        "name": "Allure Esthetic\nPlastic Surgery",
        "location": "Seattle, WA",
        "website": "allureesthetic.com",
        "google_reviews": "~1,000+ (manipulated)",
        "google_rating": "4.8 (disputed)",
        "review_sentiment": "Severely compromised — federal lawsuit for fake review manipulation; $5M judgment. Cannot be trusted.",
        "website_score": "9/10",
        "seo_blog": "Strong SEO; 'Patients' Journey' blog; ranks for 'Seattle plastic surgeon'; backed by realdrseattle® marketing.",
        "instagram": "@realdrseattle | 86K followers | Daily posts | Before/after, edu content",
        "tiktok": "Not found",
        "facebook": "6,530 likes | Regular posts",
        "insurance": "Yes — major carriers for gender-affirming surgery; no WA Medicare/Medicaid",
        "membership": "None found",
        "financing": "Alphaeon, CareCredit, United Medical Credit",
        "technology": "CO2 Laser, Vaser, J-Plasma, LED therapy, Neograft FUE/FUT | Booking: Online (proprietary)",
        "usps": "Dr. Javad Sajan brand; full surgical + medspa; transgender surgery specialist; 24/7 surgeon access; widest service range in market.",
        "weakness": "FEDERAL LAWSUIT for review fraud — massive trust deficit. Opportunity: Rani can lead with verified, authentic reviews and radical transparency.",
        "botox_price": "$14/unit (pkg: $299/30u)",
        "filler_price": "$489–$900/syringe",
        "membership_price": "N/A",
        "financing_options_short": "Alphaeon, CareCredit, UMC",
        "accepts_insurance": "Yes (limited)",
        "has_membership": "No",
        "has_blog": "Yes",
        "ig_followers": 86000,
        "google_count": 1000,
        "rating_num": 4.8,
    },
    {
        "id": 2,
        "name": "The Gallery of\nCosmetic Surgery",
        "location": "Lynnwood/Bellevue, WA",
        "website": "cosmeticsurgeryforyou.com",
        "google_reviews": "951",
        "google_rating": "4.8",
        "review_sentiment": "Highly positive. Clients praise natural results, attentive surgeons, and comprehensive post-op care. Rare negatives about pricing opacity.",
        "website_score": "8/10",
        "seo_blog": "Blog present; realdrseattle™ marketing support; procedure-specific landing pages; strong local SEO.",
        "instagram": "@galleryofcosmeticsurgery | 26K followers | High freq | Results, staff, promotions",
        "tiktok": "Not found",
        "facebook": "885 likes",
        "insurance": "No (cosmetic); limited for functional procedures",
        "membership": "None found",
        "financing": "CareCredit, Alphaeon, Prosper Healthcare Lending, United Medical Credit, in-house",
        "technology": "VASER Lipo, IPL, SmartLipo, ThermiVA, Kybella, CO2 Laser, J-Plasma | Booking: Online",
        "usps": "Board-certified surgeons (20–23+ yrs exp); JCAHO-accredited facility; 5-phase scar care; 24/7 on-call; free skin consultation; weight loss clinic.",
        "weakness": "No transparent pricing on website — deters cost-conscious clients. No membership program. Opportunity: Rani can publish clear pricing and loyalty tiers.",
        "botox_price": "Not listed",
        "filler_price": "Not listed",
        "membership_price": "N/A",
        "financing_options_short": "CareCredit, Alphaeon, Prosper, UMC",
        "accepts_insurance": "No",
        "has_membership": "No",
        "has_blog": "Yes",
        "ig_followers": 26000,
        "google_count": 951,
        "rating_num": 4.8,
    },
    {
        "id": 3,
        "name": "SkinSpirit",
        "location": "Bellevue + Seattle, WA\n(National chain)",
        "website": "skinspirit.com",
        "google_reviews": "147 (Bellevue location)",
        "google_rating": "4.7",
        "review_sentiment": "Very positive — friendly/knowledgeable staff, clean environment, effective treatments. Occasional complaints about pricing opacity and wait times.",
        "website_score": "8/10",
        "seo_blog": "Active blog (News, Wellness, People, Body, Skin categories); strong national SEO; Zenoti booking integration.",
        "instagram": "@skinspirit | 126K followers | Regular | Edu, promo, results, staff",
        "tiktok": "@skinspirit | 17.9K followers",
        "facebook": "SkinSpirit | 15,012 likes",
        "insurance": "No",
        "membership": "No formal membership; referral rewards + package deals",
        "financing": "Cherry Financing (0% APR options, 3–24 month plans)",
        "technology": "Morpheus8, Halo, BBL, CoolSculpting, CoolTone, DiamondGlow, Ellacor, Sculptra, Emsculpt | Booking: Zenoti",
        "usps": "#1 national Botox/filler provider claim; SkinSpirit University staff training; premium spa experience; broad advanced device portfolio.",
        "weakness": "Corporate chain feel — lacks local community connection. No transparent pricing. Opportunity: Rani can win on personalized, community-rooted care and published pricing.",
        "botox_price": "Not listed",
        "filler_price": "Not listed",
        "membership_price": "N/A",
        "financing_options_short": "Cherry",
        "accepts_insurance": "No",
        "has_membership": "No",
        "has_blog": "Yes",
        "ig_followers": 126000,
        "google_count": 147,
        "rating_num": 4.7,
    },
    {
        "id": 4,
        "name": "Radiance Medical Spa\n(Radiance Skincare Seattle)",
        "location": "Seattle, WA",
        "website": "radianceskincare.com",
        "google_reviews": "16",
        "google_rating": "4.9",
        "review_sentiment": "Glowing reviews praising Alicia Wright's expertise, personalized care, and beautiful results. Very small review base limits credibility.",
        "website_score": "6/10",
        "seo_blog": "No blog found. Minimal SEO. Limited keyword targeting. Acuity Scheduling booking system.",
        "instagram": "@seattle_esthetician | 239 followers | Moderate | Results, service promos",
        "tiktok": "Not found",
        "facebook": "Radiance Skincare | 255 followers",
        "insurance": "Not found",
        "membership": "None found",
        "financing": "Affirm, Afterpay, Klarna, Cherry, CareCredit",
        "technology": "Procell Microchanneling, Environ Skincare (Iontophoresis, Sonophoresis, Microcurrent), LED | Booking: Acuity Scheduling",
        "usps": "25+ years master esthetician experience; Environ Skincare specialist; Procell Microchanneling; intimate one-on-one care; science-backed product focus.",
        "weakness": "Tiny online footprint (16 reviews, 239 IG followers) — near-invisible to new clients. No blog or TikTok. Opportunity: Rani can dominate local search and social with minimal effort.",
        "botox_price": "N/A (esthetics only)",
        "filler_price": "N/A",
        "membership_price": "N/A",
        "financing_options_short": "Affirm, Afterpay, Klarna, Cherry, CareCredit",
        "accepts_insurance": "Unknown",
        "has_membership": "No",
        "has_blog": "No",
        "ig_followers": 239,
        "google_count": 16,
        "rating_num": 4.9,
    },
    {
        "id": 5,
        "name": "Vibrant Skin Bar /\nElevate Medical Center",
        "location": "Kent, WA",
        "website": "elevatemedcenter.com",
        "google_reviews": "63",
        "google_rating": "5.0",
        "review_sentiment": "Perfect 5.0 — staff (Jas Kaur, Dr. Ravneet Purewal-Nahal) praised for professionalism, personalized care, natural results. Very positive across all themes.",
        "website_score": "8/10",
        "seo_blog": "Regularly updated blog; ranks for 'Botox Kent WA' and 'medspa Kent WA'; strong local SEO foundation.",
        "instagram": "@elevatemed_ | 367 followers | Frequent | Professional, edu, before/after",
        "tiktok": "@elevatemed_ | 367 followers",
        "facebook": "Elevate Medical Center | 108 likes",
        "insurance": "Not found",
        "membership": "None found",
        "financing": "Cherry, HSA/FSA",
        "technology": "Lasers (hair removal, skin), Microneedling, PRP/PRF devices | Booking: Vagaro",
        "usps": "DNP-led holistic aesthetic + wellness approach; perfect Google rating; personalized non-invasive treatments; welcoming, inclusive environment; IV hydration + weight loss integration.",
        "weakness": "Very small social following (367 IG) and limited brand awareness outside Kent. No membership program. Opportunity: Rani can out-market them with stronger social and loyalty programs.",
        "botox_price": "$13/unit",
        "filler_price": "Lip Filler $700 | Juvederm $850",
        "membership_price": "N/A",
        "financing_options_short": "Cherry, HSA/FSA",
        "accepts_insurance": "Unknown",
        "has_membership": "No",
        "has_blog": "Yes",
        "ig_followers": 367,
        "google_count": 63,
        "rating_num": 5.0,
    },
    {
        "id": 6,
        "name": "Aesthetic Clinique /\nLa Clinique Aesthetics",
        "location": "Seattle, WA",
        "website": "lacliniqueaesthetics.com",
        "google_reviews": "58",
        "google_rating": "5.0",
        "review_sentiment": "Perfect 5.0 — praised for natural results, gentle approach, personalized plans, and welcoming LGBTQ+/women-owned environment. Some mention 'specials and low prices.'",
        "website_score": "8/10",
        "seo_blog": "Blog link present; limited content depth; basic SEO structure; no clear ranking for competitive terms.",
        "instagram": "@lacliniqueaesthetics | 250 followers | Active | Natural injectables, custom plans",
        "tiktok": "Not found",
        "facebook": "Not found",
        "insurance": "Not found",
        "membership": "None found",
        "financing": "Cherry",
        "technology": "Not found (injectables, microchanneling, PDO threads implied)",
        "usps": "Conservative anatomy-led injection philosophy; LGBTQ+ friendly; women-owned; Asian-owned; nurse injector-led; natural results focus.",
        "weakness": "No pricing transparency; tiny social footprint (250 IG followers); no Facebook presence. Opportunity: Rani can dominate in visibility and community trust.",
        "botox_price": "Not listed",
        "filler_price": "Not listed",
        "membership_price": "N/A",
        "financing_options_short": "Cherry",
        "accepts_insurance": "Unknown",
        "has_membership": "No",
        "has_blog": "Yes (limited)",
        "ig_followers": 250,
        "google_count": 58,
        "rating_num": 5.0,
    },
    {
        "id": 7,
        "name": "Pacific Dermatology &\nCosmetic Center\n(Frontier Dermatology)",
        "location": "Seattle, WA\n(38+ locations, PNW)",
        "website": "frontierdermatology.com",
        "google_reviews": "~500+ (multi-location)",
        "google_rating": "4.4",
        "review_sentiment": "Mixed — positive for specific doctors; negative for long wait times, insurance verification issues, and high costs for short appointments.",
        "website_score": "8/10",
        "seo_blog": "Active blog/news section; strong multi-location SEO; 38+ locations create massive domain authority; ranks broadly for dermatology terms.",
        "instagram": "@frontierdermatology | 4,800+ followers | Regular | Professional, informative, staff features",
        "tiktok": "@pacificdermatology | 49 followers (minimal)",
        "facebook": "Frontier Dermatology | 1,775 likes",
        "insurance": "Yes — wide range of insurance providers (Tax ID: 26-2636087); patients verify coverage directly",
        "membership": "None found",
        "financing": "Not found",
        "technology": "Morpheus8, CoolSculpting, Ultherapy, V-beam, Clear & Brilliant, Fractional CO2 (ActiveFX/TotalFX), Fraxel Dual, IPL/BBL, SkinPen | Booking: Online",
        "usps": "Largest physician-owned dermatology practice in PNW (38+ locations, 120+ providers); medical + cosmetic + surgical dermatology; insurance acceptance; clinical research + residency program.",
        "weakness": "Lowest rating (4.4) among competitors; complaints about wait times and impersonal care. Insurance billing complexity frustrates patients. Opportunity: Rani can win on speed, warmth, and simplicity.",
        "botox_price": "Not listed",
        "filler_price": "Not listed",
        "membership_price": "N/A",
        "financing_options_short": "Not found",
        "accepts_insurance": "Yes",
        "has_membership": "No",
        "has_blog": "Yes",
        "ig_followers": 4800,
        "google_count": 500,
        "rating_num": 4.4,
    },
    {
        "id": 8,
        "name": "Zen Aesthetics /\nZen's Body Sculpting",
        "location": "Federal Way, WA",
        "website": "Not found",
        "google_reviews": "Not found",
        "google_rating": "Not found",
        "review_sentiment": "No data available. Facebook presence is minimal (257 followers, infrequent posts). Essentially invisible online.",
        "website_score": "N/A — No website found",
        "seo_blog": "No website, no blog, no SEO presence. Facebook only.",
        "instagram": "Not found",
        "tiktok": "Not found",
        "facebook": "Zen's Body Sculpting | 257 followers | Infrequent | Body sculpting promos",
        "insurance": "No",
        "membership": "None found",
        "financing": "None found",
        "technology": "Laser Lipo, Ultrasonic Cavitation",
        "usps": "Budget-friendly body contouring; accessible pricing; serves all body types; local Federal Way presence.",
        "weakness": "Near-zero online presence — no website, no Google reviews, no Instagram. Essentially non-competitive digitally. Opportunity: Any digital investment by Rani will outperform this competitor.",
        "botox_price": "N/A",
        "filler_price": "N/A",
        "membership_price": "N/A",
        "financing_options_short": "None",
        "accepts_insurance": "No",
        "has_membership": "No",
        "has_blog": "No",
        "ig_followers": 0,
        "google_count": 0,
        "rating_num": 0,
    },
    {
        "id": 9,
        "name": "Spa MD /\nMedspa of Seattle",
        "location": "Bellevue, WA",
        "website": "medspaofseattle.com",
        "google_reviews": "Not found (no clear GBP)",
        "google_rating": "Not found",
        "review_sentiment": "Limited data — positive anecdotal reviews for Xeomin treatments, natural results, friendly staff. No aggregated Google review profile found.",
        "website_score": "7/10",
        "seo_blog": "Blog present with medspa treatment articles; weak local SEO (no clear Google Business Profile); needs improvement.",
        "instagram": "@medspaofseattle1 | Followers not found",
        "tiktok": "@medspaofseattle | Followers not found",
        "facebook": "Not found",
        "insurance": "No",
        "membership": "Loyalty Rewards: $39.99/month → Xeomin at $10/unit",
        "financing": "Not found",
        "technology": "Xeomin (primary), Dermal Fillers | Booking: Not found | EMR: Not found",
        "usps": "Aggressive affordability positioning ('Newer Treatments, Better Results, for Less'); Dr. Jay Park MD co-founded Botox training institution; $10/unit Xeomin for members; low-barrier entry pricing.",
        "weakness": "No Google Business Profile / aggregated reviews — invisible to patients searching locally. No Facebook presence. Opportunity: Rani can dominate local search with active GBP management.",
        "botox_price": "$10/unit (member) | Xeomin $10/unit",
        "filler_price": "$599–$699/syringe",
        "membership_price": "$39.99/month",
        "financing_options_short": "Not found",
        "accepts_insurance": "No",
        "has_membership": "Yes",
        "has_blog": "Yes",
        "ig_followers": 0,
        "google_count": 0,
        "rating_num": 0,
    },
    {
        "id": 10,
        "name": "La Belle Vie Med Spa\n(South King County equiv.)",
        "location": "Tukwila, WA",
        "website": "labelleviemed.com",
        "google_reviews": "343",
        "google_rating": "4.6",
        "review_sentiment": "Generally positive — welcoming staff, good injectable results, LGBTQIA+ friendly. Negative: some find services overpriced, lack of transparency on discounts.",
        "website_score": "7/10",
        "seo_blog": "Blog present but limited content. Appears in relevant searches but no dominant keyword rankings found.",
        "instagram": "@labelleviemed | 1,400+ followers | Irregular | Before/after, promos, staff",
        "tiktok": "Not found",
        "facebook": "La Belle Vie Medspa | 133 likes",
        "insurance": "No (medspa services)",
        "membership": "Silver: $99/mo | Gold: $175/mo | Platinum: $275/mo (tiered discounts on injectables, laser, retail + monthly service)",
        "financing": "CareCredit",
        "technology": "Morpheus8, HydraFacial, Votiva, InMode Evolve | Booking: Vagaro",
        "usps": "Tiered membership program (best in market); LGBTQIA+ inclusive; Morpheus8 + InMode Evolve technology; Tukwila/South King County presence; comprehensive non-surgical menu.",
        "weakness": "Inconsistent pricing transparency; irregular social media posting; modest Instagram following (1,400). Negative reviews cite overpricing. Opportunity: Rani can offer comparable membership tiers with clearer value.",
        "botox_price": "Not listed",
        "filler_price": "Hyperdilute Radiesse $950/syringe | Sculptra $900/vial | Kybella $800/vial",
        "membership_price": "$99–$275/month",
        "financing_options_short": "CareCredit",
        "accepts_insurance": "No",
        "has_membership": "Yes",
        "has_blog": "Yes (limited)",
        "ig_followers": 1400,
        "google_count": 343,
        "rating_num": 4.6,
    },
    {
        "id": 11,
        "name": "GS Med Spa\n(New 2025 — Renton)",
        "location": "Renton, WA",
        "website": "gsrentonmedspa.com",
        "google_reviews": "Not found",
        "google_rating": "Not found",
        "review_sentiment": "No reviews available — brand new (est. 2025). No social media presence found.",
        "website_score": "8/10",
        "seo_blog": "No blog. Copyright 2025 — too new to rank for competitive terms. Basic SEO structure in place.",
        "instagram": "Not found",
        "tiktok": "Not found",
        "facebook": "Not found",
        "insurance": "Not found",
        "membership": "None found",
        "financing": "Not found",
        "technology": "Advanced Exion Face Applicator, Exion Radiofrequency Microneedling | Booking: Not found",
        "usps": "Cutting-edge Exion RF Microneedling technology; 10+ years provider experience; luxurious tranquil environment; comprehensive aesthetic + wellness menu; newest entrant in Renton market.",
        "weakness": "Zero online reputation (no reviews, no social media). Website has no pricing. Completely unknown brand. Opportunity: Rani can establish market dominance in Renton before GS Med Spa builds awareness.",
        "botox_price": "Not listed",
        "filler_price": "Not listed",
        "membership_price": "N/A",
        "financing_options_short": "Not found",
        "accepts_insurance": "Unknown",
        "has_membership": "No",
        "has_blog": "No",
        "ig_followers": 0,
        "google_count": 0,
        "rating_num": 0,
    },
]

# ── create workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — MASTER COMPARISON MATRIX
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Master Comparison Matrix"

# --- Title row ---
ws1.merge_cells("A1:S1")
title_cell = ws1["A1"]
title_cell.value = "RANI BEAUTY CLINIC — COMPETITIVE ANALYSIS MATRIX  |  Seattle / Eastside / South King County MedSpa Market  |  March 2026"
title_cell.font = Font(name="Calibri", bold=True, size=13, color="1F3864")
title_cell.fill = fill(TITLE_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 30

# --- Column headers ---
headers = [
    "#", "Competitor", "Location", "Website",
    "Google\nReviews", "Google\nRating", "Website\nScore (1-10)",
    "Instagram\nFollowers", "TikTok\nPresence", "Facebook\nFollowers",
    "Accepts\nInsurance?", "Membership\nProgram?", "Financing\nOptions",
    "Botox\nPrice", "Filler\nPrice",
    "Key Technology\n& Devices",
    "Unique Selling\nPropositions",
    "Review Sentiment\nSummary",
    "Biggest Weakness\n(Rani's Opportunity)"
]

col_widths = [4, 22, 20, 22, 12, 10, 12, 12, 12, 12, 12, 12, 22, 14, 22, 30, 40, 35, 40]

for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws1.cell(row=2, column=col_idx, value=hdr)
    cell.font = hdr_font(size=10)
    cell.fill = fill(HEADER_BG)
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = border
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

ws1.row_dimensions[2].height = 35

# --- Data rows ---
for row_idx, c in enumerate(competitors, start=3):
    bg = ROW_ALT1 if row_idx % 2 == 0 else ROW_ALT2
    row_data = [
        c["id"],
        c["name"],
        c["location"],
        c["website"],
        c["google_reviews"],
        c["google_rating"],
        c["website_score"],
        f"{c['ig_followers']:,}" if c['ig_followers'] > 0 else "Not found",
        c["tiktok"].split("|")[0].strip() if c["tiktok"] != "Not found" else "Not found",
        c["facebook"].split("|")[1].strip() if "|" in c["facebook"] else c["facebook"],
        c["accepts_insurance"],
        c["has_membership"],
        c["financing_options_short"],
        c["botox_price"],
        c["filler_price"],
        c["technology"],
        c["usps"],
        c["review_sentiment"],
        c["weakness"],
    ]

    for col_idx, value in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = cell_font(size=9)
        cell.border = border

        # special colouring
        if col_idx == 19:   # weakness column → light red
            cell.fill = fill(ACCENT_RED)
        elif col_idx == 17:  # USPs → light green
            cell.fill = fill(ACCENT_GRN)
        elif col_idx == 18:  # sentiment → yellow
            cell.fill = fill(ACCENT_YLW)
        else:
            cell.fill = fill(bg)

        cell.alignment = wrap_align()

    ws1.row_dimensions[row_idx].height = 80

# freeze panes
ws1.freeze_panes = "E3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — PRICING COMPARISON
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Pricing Comparison")

ws2.merge_cells("A1:H1")
t2 = ws2["A1"]
t2.value = "PRICING COMPARISON — Key Services Across Competitors"
t2.font = Font(name="Calibri", bold=True, size=12, color="1F3864")
t2.fill = fill(TITLE_BG)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

pricing_headers = ["Competitor", "Location", "Botox (per unit)", "Filler (per syringe)", "Membership Price", "Financing", "Insurance?", "Pricing Transparency"]
pricing_widths  = [25, 20, 18, 28, 22, 28, 14, 22]

for col_idx, (hdr, width) in enumerate(zip(pricing_headers, pricing_widths), start=1):
    cell = ws2.cell(row=2, column=col_idx, value=hdr)
    cell.font = hdr_font(size=10)
    cell.fill = fill(SUBHDR_BG)
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = border
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

ws2.row_dimensions[2].height = 30

pricing_transparency = {
    1: "HIGH — full price list published",
    2: "LOW — no prices on website",
    3: "LOW — no prices on website",
    4: "MEDIUM — facial/sugaring prices listed",
    5: "MEDIUM — Botox + filler prices listed",
    6: "LOW — no prices on website",
    7: "LOW — no prices on website",
    8: "N/A — no website",
    9: "MEDIUM — Xeomin packages listed",
    10: "LOW — partial pricing only",
    11: "LOW — no prices on website",
}

for row_idx, c in enumerate(competitors, start=3):
    bg = ROW_ALT1 if row_idx % 2 == 0 else ROW_ALT2
    row_data = [
        c["name"].replace("\n", " "),
        c["location"].replace("\n", " "),
        c["botox_price"],
        c["filler_price"],
        c["membership_price"],
        c["financing_options_short"],
        c["accepts_insurance"],
        pricing_transparency[c["id"]],
    ]
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = cell_font(size=9)
        cell.fill = fill(bg)
        cell.border = border
        cell.alignment = wrap_align()
    ws2.row_dimensions[row_idx].height = 45

ws2.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — DIGITAL PRESENCE SCORECARD
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Digital Presence Scorecard")

ws3.merge_cells("A1:J1")
t3 = ws3["A1"]
t3.value = "DIGITAL PRESENCE SCORECARD — Social Media, SEO & Online Reputation"
t3.font = Font(name="Calibri", bold=True, size=12, color="1F3864")
t3.fill = fill(TITLE_BG)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

digital_headers = ["Competitor", "Website Score", "Has Blog?", "Instagram Followers", "TikTok", "Facebook Followers", "Google Reviews", "Google Rating", "SEO Strength", "Overall Digital Score\n(1-10)"]
digital_widths  = [25, 15, 12, 20, 20, 18, 15, 14, 25, 20]

for col_idx, (hdr, width) in enumerate(zip(digital_headers, digital_widths), start=1):
    cell = ws3.cell(row=2, column=col_idx, value=hdr)
    cell.font = hdr_font(size=10)
    cell.fill = fill(SUBHDR_BG)
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = border
    ws3.column_dimensions[get_column_letter(col_idx)].width = width

ws3.row_dimensions[2].height = 35

# Overall digital scores (manually assessed)
digital_scores = {
    1: "8/10 — Strong website + 86K IG, but review fraud taints reputation",
    2: "7/10 — Good website + 26K IG; limited social depth",
    3: "9/10 — National chain; 126K IG + TikTok + active blog",
    4: "3/10 — Minimal presence; 16 reviews; 239 IG followers",
    5: "5/10 — Good blog; very small social following",
    6: "4/10 — Small following; no Facebook; limited blog",
    7: "8/10 — 38-location SEO authority; insurance acceptance boosts visibility",
    8: "1/10 — No website, no IG, no Google presence",
    9: "3/10 — Blog exists; no GBP; social presence unclear",
    10: "5/10 — 343 reviews; 1,400 IG; limited blog",
    11: "2/10 — Brand new; no reviews; no social media",
}

seo_strength = {
    1: "Strong — realdrseattle® marketing",
    2: "Strong — realdrseattle™ marketing",
    3: "Very Strong — national brand + Zenoti",
    4: "Weak — no blog, minimal keywords",
    5: "Moderate — ranks for local Kent terms",
    6: "Weak — basic structure only",
    7: "Very Strong — 38+ locations, domain authority",
    8: "None — no website",
    9: "Weak — blog exists; no GBP",
    10: "Moderate — Vagaro + local search",
    11: "None — too new",
}

for row_idx, c in enumerate(competitors, start=3):
    bg = ROW_ALT1 if row_idx % 2 == 0 else ROW_ALT2
    tiktok_val = c["tiktok"].split("|")[0].strip() if c["tiktok"] != "Not found" else "Not found"
    fb_val = c["facebook"].split("|")[1].strip() if "|" in c["facebook"] else c["facebook"]

    row_data = [
        c["name"].replace("\n", " "),
        c["website_score"],
        c["has_blog"],
        f"{c['ig_followers']:,}" if c['ig_followers'] > 0 else "Not found",
        tiktok_val,
        fb_val,
        c["google_reviews"],
        c["google_rating"],
        seo_strength[c["id"]],
        digital_scores[c["id"]],
    ]
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = cell_font(size=9)
        cell.fill = fill(bg)
        cell.border = border
        cell.alignment = wrap_align()
    ws3.row_dimensions[row_idx].height = 45

ws3.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — STRATEGIC OPPORTUNITIES SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Strategic Opportunities")

ws4.merge_cells("A1:E1")
t4 = ws4["A1"]
t4.value = "STRATEGIC OPPORTUNITIES FOR RANI BEAUTY CLINIC"
t4.font = Font(name="Calibri", bold=True, size=13, color="1F3864")
t4.fill = fill(TITLE_BG)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

opp_headers = ["Competitor", "Primary Weakness", "Rani's Opportunity", "Priority\n(High/Med/Low)", "Recommended Action"]
opp_widths   = [25, 35, 35, 14, 45]

for col_idx, (hdr, width) in enumerate(zip(opp_headers, opp_widths), start=1):
    cell = ws4.cell(row=2, column=col_idx, value=hdr)
    cell.font = hdr_font(size=10)
    cell.fill = fill(HEADER_BG)
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = border
    ws4.column_dimensions[get_column_letter(col_idx)].width = width

ws4.row_dimensions[2].height = 30

opportunities = [
    ("Allure Esthetic", "Federal lawsuit for fake review manipulation — zero trust in online ratings", "Lead with verified, authentic patient reviews; highlight ethical transparency", "HIGH", "Launch 'Real Results' campaign with verified Google reviews; respond to every review publicly"),
    ("The Gallery of Cosmetic Surgery", "No pricing transparency; no membership program", "Publish clear pricing; launch tiered membership to capture loyalty", "HIGH", "Create a public pricing page; introduce Rani Membership tiers at $79/$149/$249/mo"),
    ("SkinSpirit", "Corporate chain — impersonal; no pricing transparency; no membership", "Win on local community connection, personalization, and published pricing", "HIGH", "Emphasize 'local, personal, transparent' in all marketing; show provider faces/stories"),
    ("Radiance Medical Spa", "Near-invisible online (16 reviews, 239 IG followers)", "Dominate local search with active GBP, blog, and social media", "MEDIUM", "Invest in Google Business Profile optimization; post 3x/week on Instagram"),
    ("Elevate Medical Center (Kent)", "Tiny social following; no membership program", "Out-market with stronger social presence and loyalty program", "MEDIUM", "Target Kent/Renton audience on Instagram/TikTok; offer referral bonuses"),
    ("La Clinique Aesthetics", "No pricing; minimal online presence (250 IG followers)", "Capture clients seeking natural results with visible pricing + strong IG", "MEDIUM", "Create 'natural results' content series; publish price ranges prominently"),
    ("Frontier Dermatology", "Lowest rating (4.4); complaints about wait times and impersonal care", "Win on speed, warmth, and patient experience", "HIGH", "Advertise same-day/next-day booking; showcase personal provider relationships"),
    ("Zen's Body Sculpting", "No website, no Google presence — digitally invisible", "Capture Federal Way body contouring clients with any digital presence", "LOW", "Add body contouring to service menu; target Federal Way on Google Ads"),
    ("Medspa of Seattle", "No Google Business Profile — invisible in local search", "Dominate local Bellevue/Seattle search with active GBP management", "MEDIUM", "Claim and optimize GBP; encourage review generation campaigns"),
    ("La Belle Vie Med Spa (Tukwila)", "Inconsistent posting; overpricing complaints; modest IG following", "Match membership tiers with clearer value + better content consistency", "HIGH", "Launch competitive membership program; post consistently 4x/week on IG"),
    ("GS Med Spa (Renton 2025)", "Zero online reputation — brand new, unknown", "Establish Renton market dominance before GS Med Spa builds awareness", "HIGH", "Geo-targeted Google Ads for Renton; SEO-optimized landing page for Renton"),
]

priority_colors = {"HIGH": "FFD7D7", "MEDIUM": "FFF2CC", "LOW": "E2EFDA"}

for row_idx, (comp, weakness, opp, priority, action) in enumerate(opportunities, start=3):
    row_data = [comp, weakness, opp, priority, action]
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font(size=9)
        cell.border = border
        cell.alignment = wrap_align()
        if col_idx == 4:
            cell.fill = fill(priority_colors.get(priority, ROW_ALT1))
            cell.font = Font(name="Calibri", bold=True, size=9)
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        else:
            bg = ROW_ALT1 if row_idx % 2 == 0 else ROW_ALT2
            cell.fill = fill(bg)
    ws4.row_dimensions[row_idx].height = 60

ws4.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SERVICES & TECHNOLOGY MATRIX
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Services & Technology")

ws5.merge_cells("A1:N1")
t5 = ws5["A1"]
t5.value = "SERVICES & TECHNOLOGY MATRIX — What Each Competitor Offers"
t5.font = Font(name="Calibri", bold=True, size=12, color="1F3864")
t5.fill = fill(TITLE_BG)
t5.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 28

service_headers = [
    "Competitor", "Botox/Neuromodulators", "Dermal Fillers", "Laser Treatments",
    "Body Contouring", "Microneedling/RF", "Facials/Peels", "Hair Restoration",
    "Surgical Procedures", "IV Therapy/Wellness", "Weight Loss", "PDO Threads",
    "Booking System", "Key Devices"
]
service_widths = [22, 14, 14, 14, 14, 16, 14, 14, 14, 14, 12, 12, 16, 30]

for col_idx, (hdr, width) in enumerate(zip(service_headers, service_widths), start=1):
    cell = ws5.cell(row=2, column=col_idx, value=hdr)
    cell.font = hdr_font(size=9)
    cell.fill = fill(SUBHDR_BG)
    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    cell.border = border
    ws5.column_dimensions[get_column_letter(col_idx)].width = width

ws5.row_dimensions[2].height = 35

services_data = [
    # name, botox, filler, laser, body_contour, microneedling, facials, hair_restore, surgical, iv_wellness, weight_loss, pdo, booking, devices
    ("Allure Esthetic", "✓", "✓", "✓ CO2, IPL", "✓ Vaser, Lipo", "✗", "✓", "✓ Neograft FUE/FUT", "✓ Full surgical", "✓ NAD+, Glutathione, Vitamin C", "✓ Ozempic/Wegovy", "✓", "Online (proprietary)", "CO2 Laser, Vaser, J-Plasma, LED, Neograft"),
    ("Gallery of Cosmetic Surgery", "✓", "✓", "✓ CO2, IPL", "✓ VASER, SmartLipo", "✗", "✓", "✓", "✓ Full surgical", "✗", "✓ Zepbound", "✗", "Online", "VASER, IPL, SmartLipo, ThermiVA, CO2, J-Plasma"),
    ("SkinSpirit", "✓", "✓", "✓ BBL, Halo, IPL", "✓ CoolSculpting, CoolTone", "✓ Morpheus8", "✓ DiamondGlow", "✗", "✗", "✗", "✗", "✓", "Zenoti", "Morpheus8, Halo, BBL, CoolSculpting, Emsculpt, DiamondGlow, Ellacor"),
    ("Radiance Medical Spa", "✗", "✗", "✗", "✗", "✓ Procell Microchanneling", "✓ Environ, LED", "✗", "✗", "✗", "✗", "✗", "Acuity Scheduling", "Procell, Environ, Microcurrent, LED"),
    ("Elevate Medical Center", "✓", "✓", "✓ Laser HR", "✗", "✓ Microneedling, PRP", "✓ Chemical Peels", "✗", "✗", "✓ IV Hydration, Nutritional Shots", "✓ Medical Weight Loss", "✗", "Vagaro", "Lasers, Microneedling, PRP/PRF devices"),
    ("La Clinique Aesthetics", "✓", "✓", "✗", "✗", "✓ Microchanneling", "✓", "✗", "✗", "✓ Vitamin Wellness", "✓ GLP-1", "✓", "Not found", "Not specified"),
    ("Frontier Dermatology", "✓", "✓", "✓ Fraxel, CO2, IPL/BBL, V-beam", "✓ CoolSculpting", "✓ SkinPen, Morpheus8", "✓ Chemical Peels", "✗", "✓ Mohs Surgery, Surgical Derm", "✗", "✗", "✗", "Online", "Morpheus8, CoolSculpting, Ultherapy, V-beam, CO2, Fraxel, IPL/BBL, SkinPen"),
    ("Zen's Body Sculpting", "✗", "✗", "✗", "✓ Laser Lipo, Cavitation", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "None", "Laser Lipo, Ultrasonic Cavitation"),
    ("Medspa of Seattle", "✓ Xeomin", "✓", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "✗", "Not found", "Xeomin, Dermal Fillers"),
    ("La Belle Vie Med Spa", "✓", "✓", "✓ IPL", "✓ InMode Evolve", "✓ Morpheus8", "✓ HydraFacial, VI Peel", "✗", "✗", "✗", "✗", "✓", "Vagaro", "Morpheus8, HydraFacial, Votiva, InMode Evolve"),
    ("GS Med Spa (Renton)", "✓ (listed)", "✓ (listed)", "✓ (listed)", "✓ (listed)", "✓ Exion RF Microneedling", "✓ (listed)", "✗", "✗", "✓ IV Vitamin Therapy", "✗", "✗", "Not found", "Exion Face Applicator, Exion RF Microneedling"),
]

for row_idx, row_data in enumerate(services_data, start=3):
    bg = ROW_ALT1 if row_idx % 2 == 0 else ROW_ALT2
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = cell_font(size=9)
        cell.border = border
        cell.alignment = wrap_align(horizontal="center" if col_idx > 1 else "left")
        if value == "✓" or value.startswith("✓"):
            cell.fill = fill(ACCENT_GRN)
        elif value == "✗":
            cell.fill = fill("F5F5F5")
        else:
            cell.fill = fill(bg)
    ws5.row_dimensions[row_idx].height = 45

ws5.freeze_panes = "B3"

# ── save ──────────────────────────────────────────────────────────────────────
output_path = "/home/ubuntu/Rani_Beauty_Clinic_Competitive_Analysis_Matrix.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
